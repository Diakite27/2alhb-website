from django.contrib import admin, messages
from django.contrib.auth.admin import UserAdmin
from django.template.response import TemplateResponse
from .models import (
    Promotion, Member, Testimonial, Event, NewsArticle,
    Partner, GalleryImage, GalleryAlbum, SiteStats, ContactMessage,
    BureauMember, JobOffer, FAQ, Activity, AssociationInfo,
    NewsletterSubscriber, MemberDocument, Notification, CotisationPayment,
)

admin.site.site_header = "Administration de 2ALHB"
admin.site.site_title = "2ALHB Admin"
admin.site.index_title = "Gestion du site"


@admin.register(Promotion)
class PromotionAdmin(admin.ModelAdmin):
    list_display = ["year", "name", "members_count"]
    search_fields = ["year", "name"]
    ordering = ["-year"]

    def members_count(self, obj):
        return obj.members.filter(is_approved=True).count()
    members_count.short_description = "Membres"


@admin.register(Member)
class MemberAdmin(UserAdmin):
    list_display = ["member_number", "username", "get_full_name", "promotion", "membership_type", "profession", "country", "is_approved"]
    list_filter = ["is_approved", "country", "promotion__year", "membership_type"]
    search_fields = ["first_name", "last_name", "email", "member_number", "promotion__year"]
    actions = ["send_cotisation_reminder", "export_members_excel"]

    add_fieldsets = (
        (None, {
            "classes": ("wide",),
            "fields": ("username", "password1", "password2"),
        }),
        ("Informations personnelles", {
            "fields": ("first_name", "last_name", "email", "phone"),
        }),
        ("Infos Alumni", {
            "fields": ("promotion", "membership_type", "cotisation_mode", "profession", "company", "city", "country"),
        }),
    )

    fieldsets = UserAdmin.fieldsets + (
        ("Infos Alumni", {
            "fields": (
                "member_number", "phone", "promotion", "membership_type", "cotisation_mode",
                "profession", "company", "city", "country", "bio", "photo",
                "linkedin", "is_approved",
            ),
        }),
    )
    readonly_fields = ["member_number"]

    @admin.action(description="📩 Envoyer un rappel de cotisation aux membres sélectionnés")
    def send_cotisation_reminder(self, request, queryset):
        from .models import Notification, CotisationPayment
        import datetime

        now = datetime.date.today()
        count = 0

        for member in queryset.filter(is_approved=True, membership_type="adherent"):
            # Trouver le dernier paiement
            last_payment = CotisationPayment.objects.filter(member=member).order_by("-paid_at").first()

            if last_payment:
                last_date = last_payment.paid_at
                # Calculer les mois impayés depuis le dernier paiement
                months_due = (now.year - last_date.year) * 12 + (now.month - last_date.month)
                if months_due <= 0:
                    continue  # À jour, pas de rappel
                period = f"depuis {last_date.strftime('%B %Y').capitalize()} ({months_due} mois)"
            else:
                period = "aucun paiement enregistré"
                months_due = "?"

            Notification.objects.create(
                recipient=member,
                title=f"Rappel de cotisation",
                message=f"Votre cotisation est en retard ({period}). Merci de régulariser votre situation dans les meilleurs délais.",
                notification_type="cotisation",
                link="/espace-membre",
            )
            count += 1

        self.message_user(request, f"✅ Rappel envoyé à {count} membre(s) en retard de cotisation.")

    @admin.action(description="📊 Exporter les membres sélectionnés en Excel")
    def export_members_excel(self, request, queryset):
        import openpyxl
        from django.http import HttpResponse

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Membres 2ALHB"

        # Header
        headers = [
            "N° Adhérent", "Nom", "Prénom", "Email", "Téléphone",
            "Promotion", "Type", "Cotisation", "Profession",
            "Entreprise", "Ville", "Pays", "Approuvé", "Date inscription",
        ]
        ws.append(headers)

        # Style header
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)

        # Data
        for member in queryset.select_related("promotion"):
            ws.append([
                member.member_number or "",
                member.last_name,
                member.first_name,
                member.email,
                member.phone,
                str(member.promotion.year) if member.promotion else "",
                member.get_membership_type_display(),
                member.get_cotisation_mode_display() if member.cotisation_mode else "",
                member.profession,
                member.company,
                member.city,
                member.country,
                "Oui" if member.is_approved else "Non",
                member.created_at.strftime("%d/%m/%Y") if member.created_at else "",
            ])

        # Auto-width columns
        for col in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 40)

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="membres_2alhb.xlsx"'
        wb.save(response)
        return response


@admin.register(Testimonial)
class TestimonialAdmin(admin.ModelAdmin):
    list_display = ["member", "is_featured", "created_at"]
    list_filter = ["is_featured"]


@admin.register(Event)
class EventAdmin(admin.ModelAdmin):
    list_display = ["title", "date", "category", "location", "is_featured", "is_published"]
    list_filter = ["is_published", "is_featured", "category"]


@admin.register(NewsArticle)
class NewsArticleAdmin(admin.ModelAdmin):
    list_display = ["title", "author", "is_published", "published_at"]
    list_filter = ["is_published"]
    prepopulated_fields = {"slug": ("title",)}


@admin.register(Partner)
class PartnerAdmin(admin.ModelAdmin):
    list_display = ["name", "website", "order"]
    list_editable = ["order"]


@admin.register(GalleryImage)
class GalleryImageAdmin(admin.ModelAdmin):
    list_display = ["title", "album", "event", "created_at"]
    list_filter = ["album", "event"]


@admin.register(SiteStats)
class SiteStatsAdmin(admin.ModelAdmin):
    list_display = ["members_count", "countries_count", "promotions_count", "insertion_rate"]

    def has_add_permission(self, request):
        return not SiteStats.objects.exists()

    def has_delete_permission(self, request, obj=None):
        return False


@admin.register(ContactMessage)
class ContactMessageAdmin(admin.ModelAdmin):
    list_display = ["name", "email", "subject", "is_read", "created_at"]
    list_filter = ["is_read"]
    readonly_fields = ["name", "email", "subject", "message", "created_at"]


@admin.register(BureauMember)
class BureauMemberAdmin(admin.ModelAdmin):
    list_display = ["get_display_name", "role", "category", "order"]
    list_filter = ["category"]
    list_editable = ["order"]

    @admin.display(description="Nom")
    def get_display_name(self, obj):
        return obj.display_name


@admin.register(JobOffer)
class JobOfferAdmin(admin.ModelAdmin):
    list_display = ["title", "company", "job_type", "sector", "is_active", "created_at"]
    list_filter = ["is_active", "job_type", "sector"]
    search_fields = ["title", "company", "description"]


@admin.register(FAQ)
class FAQAdmin(admin.ModelAdmin):
    list_display = ["question", "order", "is_published"]
    list_filter = ["is_published"]
    list_editable = ["order", "is_published"]


@admin.register(Activity)
class ActivityAdmin(admin.ModelAdmin):
    list_display = ["title", "quarter", "year", "status", "order"]
    list_filter = ["quarter", "year", "status"]
    list_editable = ["status", "order"]


@admin.register(AssociationInfo)
class AssociationInfoAdmin(admin.ModelAdmin):
    list_display = ["name", "email", "phone"]
    fieldsets = (
        ("Informations générales", {
            "fields": ("name", "full_name", "slogan", "email", "phone", "address", "whatsapp"),
        }),
        ("Réseaux sociaux", {
            "fields": ("facebook_url", "linkedin_url"),
        }),
        ("Tarifs", {
            "fields": ("adhesion_fee", "monthly_fee", "annual_fee"),
        }),
        ("Email d'accueil (personnalisable)", {
            "fields": ("welcome_email_subject", "welcome_email_body"),
            "description": "Ce message est envoyé en complément du mail technique contenant les identifiants de connexion. Variables : {prenom}, {nom}, {promotion}, {type_membre}",
        }),
    )

    def has_add_permission(self, request):
        return not AssociationInfo.objects.exists()

    def has_delete_permission(self, request, obj=None):
        return False


@admin.register(NewsletterSubscriber)
class NewsletterSubscriberAdmin(admin.ModelAdmin):
    list_display = ["email", "is_active", "subscribed_at"]
    list_filter = ["is_active"]
    search_fields = ["email"]
    actions = ["send_custom_newsletter"]

    @admin.action(description="📧 Envoyer une newsletter aux abonnés sélectionnés")
    def send_custom_newsletter(self, request, queryset):
        """Action admin pour envoyer une newsletter personnalisée."""
        if "apply" in request.POST:
            subject = request.POST.get("subject", "").strip()
            body = request.POST.get("body", "").strip()

            if not subject or not body:
                self.message_user(request, "❌ L'objet et le contenu sont obligatoires.", messages.ERROR)
                return

            active_emails = list(queryset.filter(is_active=True).values_list("email", flat=True))
            if not active_emails:
                self.message_user(request, "❌ Aucun abonné actif dans la sélection.", messages.WARNING)
                return

            from .emails import send_newsletter_custom

            sent_count = send_newsletter_custom(subject, body, recipient_emails=active_emails)

            self.message_user(
                request,
                f"✅ Newsletter envoyée avec succès à {sent_count} abonné(s).",
                messages.SUCCESS,
            )
            return

        # Afficher le formulaire intermédiaire
        return TemplateResponse(request, "admin/newsletter_send_form.html", {
            "title": "Envoyer une newsletter",
            "queryset": queryset,
            "opts": self.model._meta,
            "action_checkbox_name": admin.helpers.ACTION_CHECKBOX_NAME,
            "subscribers_count": queryset.filter(is_active=True).count(),
        })


class GalleryImageInline(admin.TabularInline):
    model = GalleryImage
    extra = 1
    fields = ["title", "image", "caption"]


@admin.register(GalleryAlbum)
class GalleryAlbumAdmin(admin.ModelAdmin):
    list_display = ["title", "event", "photos_count", "is_published", "created_at"]
    list_filter = ["is_published"]
    inlines = [GalleryImageInline]

    def photos_count(self, obj):
        return obj.images.count()
    photos_count.short_description = "Photos"


@admin.register(MemberDocument)
class MemberDocumentAdmin(admin.ModelAdmin):
    list_display = ["title", "category", "is_adherent_only", "published_at"]
    list_filter = ["category", "is_adherent_only"]
    search_fields = ["title"]
    date_hierarchy = "published_at"


@admin.register(Notification)
class NotificationAdmin(admin.ModelAdmin):
    list_display = ["title", "recipient", "notification_type", "is_read", "created_at"]
    list_filter = ["notification_type", "is_read"]
    search_fields = ["title", "recipient__first_name", "recipient__last_name"]


@admin.register(CotisationPayment)
class CotisationPaymentAdmin(admin.ModelAdmin):
    list_display = ["member", "category", "amount", "period_label", "payment_method", "paid_at"]
    list_filter = ["category", "payment_method", "paid_at"]
    search_fields = ["member__first_name", "member__last_name", "period_label", "reference"]
    date_hierarchy = "paid_at"
    actions = ["export_payments_excel"]

    @admin.action(description="📊 Exporter les paiements sélectionnés en Excel")
    def export_payments_excel(self, request, queryset):
        import openpyxl
        from django.http import HttpResponse

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Paiements 2ALHB"

        headers = [
            "N° Adhérent", "Nom", "Prénom", "Catégorie", "Montant (FCFA)",
            "Période", "Mode de paiement", "Référence", "Date de paiement",
        ]
        ws.append(headers)

        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)

        for p in queryset.select_related("member"):
            ws.append([
                p.member.member_number or "",
                p.member.last_name,
                p.member.first_name,
                p.get_category_display(),
                p.amount,
                p.period_label,
                p.payment_method,
                p.reference,
                p.paid_at.strftime("%d/%m/%Y") if p.paid_at else "",
            ])

        for col in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 40)

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="paiements_2alhb.xlsx"'
        wb.save(response)
        return response
